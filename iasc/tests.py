import datetime
import io
import json
from random import randrange
from zipfile import ZipFile

from django.utils import timezone
from django.contrib.auth.models import User
from django.test import TestCase
from django.core.exceptions import ObjectDoesNotExist
from django.core.files.uploadedfile import SimpleUploadedFile
from urllib import parse

from django.utils.text import slugify

from iasc.models import Institution, Discipline, Participant, ActiveLink, Survey, Result
from frontend.tests import HTTPTestCase

from openpyxl import Workbook, load_workbook


class ViewsTestCase(HTTPTestCase):
    """
    Test Views using Integration tests
    """

    def __init__(self, methodName: str = ...):
        super().__init__(methodName)
        self.test_data = [
            [
                ("Ariadne", "ar.kimb@durham.test"),
                ("Allan", "allan_hy@harvard.test"),
                ("Zonda", "zond-yor@durham.test"),
                ("Iestyn", "ie_molli@cambridge.test"),
                ("Hart", "ha.turco@newcastle.test"),
            ],
            [
                ("Haldis", "haldi-otoole@newcastle.test"),
                ("Cadmus", "ca.weym@harvard.test"),
                ("Jewel", "jewe.atte@durham.test"),
                ("Nimesh", "nimes_luong@yale.test"),
                ("Avery", "aver_weymo@cambridge.test"),
            ],
            [
                ("Cronan", "crona.nimmo@newcastle.test"),
                ("Nora", "nora-huf@harvard.test"),
                ("Kalil", "kalil_easter@cambridge.test"),
                ("Archidamus", "archidam.ma@harvard.test"),
                ("Betsey", "be-bauma@yale.test"),
            ],
        ]

        self.flat_test_data = [item for sublist in self.test_data for item in sublist]

        self.mimetypes = {
            "json": "application/json",
            "zip": "application/zip",
            "xlsx": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        }

        self.test_question = "Were the dish and the spoon complicit in the cow's crime?"
        self.test_institution = "Test University"
        self.sheet_header = ["First Name", "E-mail Address", "Unique Link"]

        self.links = None
        self.survey_id = None
        self.l3c_survey_id = None
        self.institution_id = None

        self.username = "testuser"
        self.password = "12345"

    def setUp(self):
        super(ViewsTestCase, self).setUp()

        user = User.objects.create(username=self.username)
        user.set_password(self.password)
        user.save()

        logged_in = self.client.login(username=self.username, password=self.password)
        self.assertTrue(logged_in)

    def helper_create_workbook(self, data):
        wb = Workbook()
        sheet1 = wb.active
        sheet1.title = "Physics"
        wb.create_sheet("Biology")
        wb.create_sheet("Chemistry")

        for i, sheet in enumerate(wb):
            sheet["A1"] = self.sheet_header[0]
            sheet["B1"] = self.sheet_header[1]

            for row, (name, email) in enumerate(data[i], start=2):
                sheet[f"A{row}"] = name
                sheet[f"B{row}"] = email

        with io.BytesIO() as buffer:
            wb.save(buffer)
            buffer.seek(0)
            return buffer.read()

    @staticmethod
    def helper_get_xls_data(data):
        read = io.BytesIO(data)
        wb = load_workbook(read)
        ws = wb.active

        # Return table as list of lists, where each sub-list is a row:
        return [[cell.value for cell in row] for row in ws.rows]

    @staticmethod
    def helper_zip_get_files(data):
        read = io.BytesIO(data)
        input_zip = ZipFile(read)
        return input_zip.namelist(), input_zip

    def helper_test_zipped_sheets(
        self, route: str, slug: str, prefix: str, headers: [], file_id: int = None
    ):
        resp = self.GET(
            f"/api{route}",
            mimetype=self.mimetypes["zip"],
            status=200,
            startswith=b"PK",
        )

        names, input_zip = self.helper_zip_get_files(resp.content)
        filename = slugify(slug, allow_unicode=True)
        id_part = file_id if file_id is not None else self.survey_id
        expected_filename = f"{prefix}-{id_part}-{filename}.xlsx"
        self.assertEqual(names[0], expected_filename)

        xls = input_zip.read(names[0])
        # Excel files are zip files too:
        self.assertTrue(xls.startswith(b"PK"))

        # Look for xlsx-specific strings:
        self.assertTrue(b"docProps/app.xml" in xls)
        self.assertTrue(b"/_rels/workbook.xml.rels" in xls)

        # Check XLS data:
        data = self.helper_get_xls_data(xls)
        self.assertEqual(data.pop(0), headers)
        self.assertEqual(len(data), len(self.flat_test_data))

    def test_api_routes(self):
        """
        API routes Integration Test:
          /participants/upload/
          /participants/

          /institutions/
          /disciplines/

          /survey/create/
          /survey/survey_id/institutions/
          /survey/close/
          /survey/results/
          /survey/

          /vote/

          /links/xls/
          /links/zip/
          /links/

          /result/xls/
          /result/zip/
          /result/                         ️

          /user/


        These are within a wrapper method as django rolls back the database state in-between
        https://docs.djangoproject.com/en/4.2/topics/testing/overview/#the-test-database
        """

        def test_01_participants_upload():
            """
            Test the "Upload Participants" route by uploading an Excel file
            /participants/upload/
            /participants/
            """

            stream = self.helper_create_workbook(self.test_data)

            upload = SimpleUploadedFile(
                "file.xlsx",
                stream,
                content_type="application/vnd.ms-excel",
            )
            self.POST(
                "/api/participants/upload/",
                {
                    "file": upload,
                    "institution": self.test_institution,
                    "create_institutions": True,
                    "create_disciplines": True,
                },
                status=200,
                mimetype=self.mimetypes["json"],
                contains="File uploaded",
            )

            self.institution_id = Institution.objects.get(name=self.test_institution).id
            self.assertEqual(Participant.objects.count(), len(self.flat_test_data))

            for row in self.flat_test_data:
                self.assertEqual(
                    Participant.objects.filter(email=row[1]).get().name,
                    row[0],
                )

            # Test participant JSON retrieval
            resp = self.GET(
                "/api/participants/", mimetype=self.mimetypes["json"], startswith=b"{"
            )

            self.assertEqual(resp["count"], len(self.flat_test_data))

        def test_02_create_survey():
            """
            Create a survey in the database
            /survey/create/
            """

            resp = self.POST(
                "/api/survey/create/",
                {
                    "question": self.test_question,
                    "expiry": "2023-06-23T23:00",
                    "active": "True",
                    "create_active_links": "True",
                },
                mimetype=self.mimetypes["json"],
            )

            self.assertEqual(resp["status"], "success")
            self.assertEqual(resp["message"], "Survey created.")
            self.assertEqual(Survey.objects.count(), 1)
            self.assertEqual(ActiveLink.objects.count(), len(self.flat_test_data))

        def test_03_download_links():
            """
            Test Active Surveys API Route
            /survey/?active=true
            /survey/survey_id/institutions/
            /links/xls/
            """

            resp_surveys = self.GET(
                "/api/survey/?active=true",
                mimetype=self.mimetypes["json"],
                startswith=b"{",
            )

            self.assertEqual(resp_surveys["results"][0]["question"], self.test_question)

            self.survey_id = resp_surveys["results"][0]["id"]
            resp_institutions = self.GET(
                f"/api/survey/{self.survey_id}/institutions/",
                mimetype=self.mimetypes["json"],
                startswith=b"{",
            )

            self.assertEqual(
                resp_institutions["results"][0]["name"], self.test_institution
            )

            # Test without ?survey=n, should raise 400
            resp = self.GET(
                "/api/links/xls/",
                mimetype=self.mimetypes["json"],
                status=400,
                startswith=None,
            )

            self.assertTrue("?survey=n" in resp["message"])

            resp = self.GET(
                f"/api/links/xls/?survey={self.survey_id}",
                mimetype=self.mimetypes["xlsx"],
                status=200,
                # Look for Zip header at the beginning of the binary response
                startswith=b"PK",
            )

            data = self.helper_get_xls_data(resp.content)

            # Check that the Excel sheet returned contains the correct header
            self.assertEqual(data.pop(0), self.sheet_header)
            # Check that the Excel sheet contains the correct number of rows
            self.assertEqual(len(data), len(self.flat_test_data))

            # Store links in variable for later use
            self.links = data

        def test_04_zip_links():
            """
            Test Zip route for links download
            /links/zip/
            """

            self.helper_test_zipped_sheets(
                f"/links/zip/?survey={self.survey_id}",
                self.test_institution,
                "CSCOPE",
                self.sheet_header,
                file_id=self.institution_id,
            )

        def test_05_vote():
            """
            Test the ability to vote using the ActiveLinks returned by the API
            /vote/
            /links/
            """

            count = 0
            for _a, _b, link in self.links:
                path = parse.urlparse(link)
                params = parse.parse_qs(path.query)

                resp = self.POST(
                    "/api/vote/",
                    {"unique_id": params["unique_id"][0], "vote": randrange(1, 6)},
                    status=200,
                    mimetype=self.mimetypes["json"],
                    contains="success",
                )

                count += 1

                self.assertEqual(resp["status"], "success")
                self.assertEqual(resp["message"], "Voted")

                # Verify that voting is incrementing the vote count on the survey
                result = self.GET(
                    f"/api/survey/{self.survey_id}/",
                    status=200,
                    mimetype=self.mimetypes["json"],
                    startswith=b"{",
                    contains="voted",
                )

                self.assertEqual(result["voted"], count)

            resp = self.GET(
                "/api/links/",
                status=200,
                mimetype=self.mimetypes["json"],
                startswith=b"{",
                contains="results",
            )

            # Check that all voting links are used up
            self.assertEqual(resp["count"], 0)

        def test_06_close():
            """
            Test survey closure
            /survey/close/
            """

            resp = self.POST(
                "/api/survey/close/",
                {"survey": self.survey_id},
                status=200,
                mimetype=self.mimetypes["json"],
                contains="success",
            )

            self.assertTrue("Closed survey" in resp["message"])

        def test_07_results():
            """
            Test results download
            /result/
            /result/xls/
            /survey/results/
            """

            resp = self.GET(
                f"/api/result/",
                status=200,
                mimetype=self.mimetypes["json"],
                startswith=b"{",
            )

            self.assertEqual(resp["count"], len(self.flat_test_data))

            resp = self.GET(
                f"/api/result/xls/?survey={self.survey_id}",
                status=200,
                mimetype=self.mimetypes["xlsx"],
                startswith=b"PK",
            )

            data = self.helper_get_xls_data(resp.content)

            # Check that the Excel sheet returned contains the correct header and no. of rows
            self.assertEqual(
                data.pop(0), ["vote", "institution", "discipline", "added"]
            )
            self.assertEqual(len(data), len(self.flat_test_data))

            resp = self.GET(
                "/api/survey/results/",
                status=200,
                mimetype=self.mimetypes["json"],
                startswith=b"{",
            )

            self.assertEqual(resp["count"], 1)
            results = resp["results"][0]
            self.assertEqual(results["question"], self.test_question)
            self.assertEqual(results["count"], len(self.flat_test_data))
            self.assertEqual(results["kind"], "LI")
            self.assertEqual(results["active"], "False")

        def test_08_zip_results():
            """
            Test route for zip results download
            /result/zip/
            """

            self.helper_test_zipped_sheets(
                f"/result/zip/?survey={self.survey_id}",
                self.test_question,
                "Results",
                ["survey", "vote", "institution", "discipline", "added"],
            )

            self.helper_test_zipped_sheets(
                f"/result/zip/",
                self.test_question,
                "Results",
                ["survey", "vote", "institution", "discipline", "added"],
            )

        def test_09_institution_discipline():
            """
            Test these routes:
            /institutions/
            /disciplines/
            @return:
            """

            self.GET(
                "/api/institutions/",
                status=200,
                mimetype=self.mimetypes["json"],
                startswith=b"[",
                contains=self.test_institution,
            )

            resp = self.GET(
                "/api/disciplines/",
                status=200,
                mimetype=self.mimetypes["json"],
                startswith=b"[",
                contains=None,
            )

            # Check for the three added disciplines
            self.assertEqual(len(resp), 3)

        def test_10_user():
            """
            Test logged in user detail route
            /user/
            """

            resp = self.GET(
                "/api/user/",
                status=200,
                mimetype=self.mimetypes["json"],
                startswith=None,
                contains=None,
            )[0]

            self.assertTrue("username" in resp.keys())
            self.assertTrue("first_name" in resp.keys())
            self.assertEqual(resp["username"], self.username)

        def test_11_create_survey_with_institution():
            """
            Test creating survey with institution parameter
            """
            with_institution = "Different University"
            institution = Institution.objects.create(name=with_institution)
            discipline = Discipline.objects.get(name="Physics")
            Participant.objects.create(
                email="test2@mytest.test",
                name="TestParticipant",
                institution=institution,
                discipline=discipline,
            )

            resp = self.POST(
                "/api/survey/create/",
                {
                    "question": "Test Create Survey With Institution",
                    "expiry": "2023-06-24T23:00",
                    "active": "True",
                    "institution": institution.id,
                    "create_active_links": "True",
                },
                mimetype=self.mimetypes["json"],
            )

            self.assertEqual(resp["status"], "success")
            self.assertEqual(resp["message"], "Survey created.")
            self.assertEqual(Survey.objects.count(), 2)
            self.assertEqual(ActiveLink.objects.count(), 1)

        def test_12_create_l3c_survey():
            """
            Test creating an L3C (3 Likert + 1 checkbox) survey
            /survey/create/
            /survey/
            """
            import json

            questions = [
                "Statement one for L3C survey",
                "Statement two for L3C survey",
                "Statement three for L3C survey",
                "I have relevant expertise in L3C",
            ]

            resp = self.POST(
                "/api/survey/create/",
                {
                    "question": "L3C parent question",
                    "questions": json.dumps(questions),
                    "expiry": "2030-01-01T00:00",
                    "active": "True",
                    "kind": "L3C",
                    "create_active_links": "True",
                },
                mimetype=self.mimetypes["json"],
            )

            self.assertEqual(resp["status"], "success")
            survey = Survey.objects.filter(kind="L3C").get()
            self.assertEqual(survey.kind, "L3C")
            self.assertEqual(survey.questions, questions)

            # Verify questions are returned by the survey list endpoint
            resp_survey = self.GET(
                f"/api/survey/{survey.id}/",
                status=200,
                mimetype=self.mimetypes["json"],
                startswith=b"{",
            )
            self.assertEqual(resp_survey["kind"], "L3C")
            self.assertEqual(resp_survey["questions"], questions)

            self.l3c_survey_id = survey.id

        def test_13_vote_l3c():
            """
            Test submitting a dict vote for an L3C survey, and verify vote_counts aggregation
            /vote/
            /survey/results/
            """
            import json

            # Get a voting link for the L3C survey
            resp_links = self.GET(
                f"/api/links/?survey={self.l3c_survey_id}",
                status=200,
                mimetype=self.mimetypes["json"],
                startswith=b"{",
            )
            self.assertGreater(resp_links["count"], 0)

            link_url = resp_links["results"][0]["hyperlink"]
            params = parse.parse_qs(parse.urlparse(link_url).query)
            unique_id = params["unique_id"][0]

            vote = {"0": 3, "1": 4, "2": 2, "3": True}

            resp = self.POST(
                "/api/vote/",
                {"unique_id": unique_id, "vote": json.dumps(vote)},
                status=200,
                mimetype=self.mimetypes["json"],
                contains="success",
            )
            self.assertEqual(resp["status"], "success")

            # Verify the result was stored as a dict
            result = Result.objects.filter(survey_id=self.l3c_survey_id).get()
            self.assertIsInstance(result.vote, dict)
            self.assertEqual(result.vote["0"], 3)
            self.assertEqual(result.vote["3"], True)

            # Verify vote_counts aggregation is nested for L3C
            resp_results = self.GET(
                "/api/survey/results/",
                status=200,
                mimetype=self.mimetypes["json"],
                startswith=b"{",
            )
            l3c_result = next(r for r in resp_results["results"] if r["kind"] == "L3C")
            self.assertIn("0", l3c_result["vote_counts"])
            self.assertIsInstance(l3c_result["vote_counts"]["0"], dict)
            self.assertEqual(l3c_result["vote_counts"]["0"]["3"], 1)
            self.assertEqual(l3c_result["vote_counts"]["3"]["True"], 1)

        def test_14_create_l2e_survey():
            """
            Test creating an L2E (2 Likert + 1 checkbox) survey
            /survey/create/
            /survey/
            """
            import json

            questions = [
                "Statement one for L2E survey",
                "Statement two for L2E survey",
                "I have relevant expertise in L2E",
            ]

            resp = self.POST(
                "/api/survey/create/",
                {
                    "question": "L2E parent question",
                    "questions": json.dumps(questions),
                    "expiry": "2030-01-01T00:00",
                    "active": "True",
                    "kind": "L2E",
                    "create_active_links": "True",
                },
                mimetype=self.mimetypes["json"],
            )

            self.assertEqual(resp["status"], "success")
            survey = Survey.objects.filter(kind="L2E").get()
            self.assertEqual(survey.kind, "L2E")
            self.assertEqual(survey.questions, questions)

            resp_survey = self.GET(
                f"/api/survey/{survey.id}/",
                status=200,
                mimetype=self.mimetypes["json"],
                startswith=b"{",
            )
            self.assertEqual(resp_survey["kind"], "L2E")
            self.assertEqual(resp_survey["questions"], questions)

            self.l2e_survey_id = survey.id

        def test_15_vote_l2e():
            """
            Test submitting a dict vote for an L2E survey, and verify vote_counts aggregation
            /vote/
            /survey/results/
            """
            import json

            resp_links = self.GET(
                f"/api/links/?survey={self.l2e_survey_id}",
                status=200,
                mimetype=self.mimetypes["json"],
                startswith=b"{",
            )
            self.assertGreater(resp_links["count"], 0)

            link_url = resp_links["results"][0]["hyperlink"]
            params = parse.parse_qs(parse.urlparse(link_url).query)
            unique_id = params["unique_id"][0]

            vote = {"0": 2, "1": 4, "2": False}

            resp = self.POST(
                "/api/vote/",
                {"unique_id": unique_id, "vote": json.dumps(vote)},
                status=200,
                mimetype=self.mimetypes["json"],
                contains="success",
            )
            self.assertEqual(resp["status"], "success")

            result = Result.objects.filter(survey_id=self.l2e_survey_id).get()
            self.assertIsInstance(result.vote, dict)
            self.assertEqual(result.vote["0"], 2)
            self.assertEqual(result.vote["2"], False)

            resp_results = self.GET(
                "/api/survey/results/",
                status=200,
                mimetype=self.mimetypes["json"],
                startswith=b"{",
            )
            l2e_result = next(r for r in resp_results["results"] if r["kind"] == "L2E")
            self.assertIn("0", l2e_result["vote_counts"])
            self.assertIsInstance(l2e_result["vote_counts"]["0"], dict)
            self.assertEqual(l2e_result["vote_counts"]["0"]["2"], 1)
            self.assertEqual(l2e_result["vote_counts"]["2"]["False"], 1)

        def test_16_create_li3_survey():
            """
            Test creating an LI3 (3 Likert, no expertise) survey
            /survey/create/
            /survey/
            """
            import json

            questions = [
                "Statement one for LI3 survey",
                "Statement two for LI3 survey",
                "Statement three for LI3 survey",
            ]

            resp = self.POST(
                "/api/survey/create/",
                {
                    "question": "LI3 parent question",
                    "questions": json.dumps(questions),
                    "expiry": "2030-01-01T00:00",
                    "active": "True",
                    "kind": "LI3",
                    "create_active_links": "True",
                },
                mimetype=self.mimetypes["json"],
            )

            self.assertEqual(resp["status"], "success")
            survey = Survey.objects.filter(kind="LI3").get()
            self.assertEqual(survey.kind, "LI3")
            self.assertEqual(survey.questions, questions)

            resp_survey = self.GET(
                f"/api/survey/{survey.id}/",
                status=200,
                mimetype=self.mimetypes["json"],
                startswith=b"{",
            )
            self.assertEqual(resp_survey["kind"], "LI3")
            self.assertEqual(resp_survey["questions"], questions)

            self.li3_survey_id = survey.id

        def test_17_vote_li3():
            """
            Test submitting a dict vote for an LI3 survey, and verify vote_counts aggregation
            /vote/
            /survey/results/
            """
            import json

            resp_links = self.GET(
                f"/api/links/?survey={self.li3_survey_id}",
                status=200,
                mimetype=self.mimetypes["json"],
                startswith=b"{",
            )
            self.assertGreater(resp_links["count"], 0)

            link_url = resp_links["results"][0]["hyperlink"]
            params = parse.parse_qs(parse.urlparse(link_url).query)
            unique_id = params["unique_id"][0]

            vote = {"0": 5, "1": 3, "2": 1}

            resp = self.POST(
                "/api/vote/",
                {"unique_id": unique_id, "vote": json.dumps(vote)},
                status=200,
                mimetype=self.mimetypes["json"],
                contains="success",
            )
            self.assertEqual(resp["status"], "success")

            result = Result.objects.filter(survey_id=self.li3_survey_id).get()
            self.assertIsInstance(result.vote, dict)
            self.assertEqual(result.vote["0"], 5)
            self.assertEqual(result.vote["1"], 3)
            self.assertEqual(result.vote["2"], 1)

            resp_results = self.GET(
                "/api/survey/results/",
                status=200,
                mimetype=self.mimetypes["json"],
                startswith=b"{",
            )
            li3_result = next(r for r in resp_results["results"] if r["kind"] == "LI3")
            self.assertIn("0", li3_result["vote_counts"])
            self.assertIsInstance(li3_result["vote_counts"]["0"], dict)
            self.assertEqual(li3_result["vote_counts"]["0"]["5"], 1)
            self.assertEqual(li3_result["vote_counts"]["1"]["3"], 1)
            self.assertEqual(li3_result["vote_counts"]["2"]["1"], 1)
            self.assertNotIn("expertise", li3_result["vote_counts"])

        def test_18_xls_multi_survey_results():
            """
            Regression test: XLS export for multi-question surveys must expand the
            vote dict into individual columns (vote_0, vote_1, vote_2, ...)
            rather than leaving a single blank 'vote' cell.
            /result/xls/
            """

            # L3C survey: vote = {"0": 3, "1": 4, "2": 2, "3": True}
            resp = self.GET(
                f"/api/result/xls/?survey={self.l3c_survey_id}",
                status=200,
                mimetype=self.mimetypes["xlsx"],
                startswith=b"PK",
            )
            data = self.helper_get_xls_data(resp.content)
            headers = data.pop(0)

            self.assertIn("vote_0", headers)
            self.assertIn("vote_1", headers)
            self.assertIn("vote_2", headers)
            self.assertIn("vote_3", headers)
            self.assertNotIn("vote", headers)

            row = dict(zip(headers, data[0]))
            self.assertEqual(row["vote_0"], 3)
            self.assertEqual(row["vote_1"], 4)
            self.assertEqual(row["vote_2"], 2)
            self.assertEqual(row["vote_3"], True)

            # L2E survey: vote = {"0": 2, "1": 4, "2": False}
            resp = self.GET(
                f"/api/result/xls/?survey={self.l2e_survey_id}",
                status=200,
                mimetype=self.mimetypes["xlsx"],
                startswith=b"PK",
            )
            data = self.helper_get_xls_data(resp.content)
            headers = data.pop(0)

            self.assertIn("vote_0", headers)
            self.assertIn("vote_1", headers)
            self.assertIn("vote_2", headers)
            self.assertNotIn("vote", headers)

            row = dict(zip(headers, data[0]))
            self.assertEqual(row["vote_0"], 2)
            self.assertEqual(row["vote_1"], 4)
            self.assertEqual(row["vote_2"], False)

            # LI3 survey: vote = {"0": 5, "1": 3, "2": 1} — no expertise column
            resp = self.GET(
                f"/api/result/xls/?survey={self.li3_survey_id}",
                status=200,
                mimetype=self.mimetypes["xlsx"],
                startswith=b"PK",
            )
            data = self.helper_get_xls_data(resp.content)
            headers = data.pop(0)

            self.assertIn("vote_0", headers)
            self.assertIn("vote_1", headers)
            self.assertIn("vote_2", headers)
            self.assertNotIn("vote_expertise", headers)
            self.assertNotIn("vote", headers)

            row = dict(zip(headers, data[0]))
            self.assertEqual(row["vote_0"], 5)
            self.assertEqual(row["vote_1"], 3)
            self.assertEqual(row["vote_2"], 1)

        #
        # Run all the integration tests defined within this function:
        print(f"\n{len(dir())-1} Integration tests found")

        for fn in dir():
            if fn == "self":
                continue
            print(f"Running {fn}")
            ret = locals()[fn]()


class DatabaseModelTestCase(TestCase):
    def setUp(self):
        super().setUp()
        self.institution = Institution.objects.create(
            name="Durham University", country="GB"
        )
        self.discipline = Discipline.objects.create(name="Research Software Engineer")
        self.participant = Participant.objects.create(
            name="Jane Doe",
            title="Dr",
            email="jane.doe@example.invalid",
            institution=self.institution,
            discipline=self.discipline,
        )

        self.survey = Survey.objects.create(
            question="What is the causative factor in Bovine Lunar Springing (BLS)?",
            active=True,
            kind="LI",
            expiry=datetime.datetime(2099, 1, 1, 0, 0),
            participants=1,
            voted=0,
        )

        ActiveLink.objects.create(participant=self.participant, survey=self.survey)

    def test_database_models(self):
        link = ActiveLink.objects.get()
        uid = str(link.unique_link)
        vote = str(5)
        link.vote(vote)

        # After voting, the unique link *must not exist* in the database
        # Check that we raise ObjectDoesNotExist looking up the uid:
        self.assertRaises(ObjectDoesNotExist, ActiveLink.objects.get, unique_link=uid)

    def test_l3c_survey_creation(self):
        """L3C survey stores questions list and correct kind."""
        questions = ["Statement A", "Statement B", "Statement C"]
        survey = Survey.objects.create(
            question="L3C parent",
            questions=questions,
            active=True,
            kind="L3C",
            expiry=datetime.datetime(2099, 1, 1, 0, 0),
            participants=1,
            voted=0,
        )
        saved = Survey.objects.get(pk=survey.pk)
        self.assertEqual(saved.kind, "L3C")
        self.assertEqual(saved.questions, questions)

    def test_l3c_vote_stored_as_dict(self):
        """Voting on an L3C survey stores a dict in Result.vote."""
        survey = Survey.objects.create(
            question="L3C vote test",
            questions=["S1", "S2", "S3"],
            active=True,
            kind="L3C",
            expiry=datetime.datetime(2099, 1, 1, 0, 0),
            participants=1,
            voted=0,
        )
        link = ActiveLink.objects.create(participant=self.participant, survey=survey)
        vote = {"0": 3, "1": 4, "2": 2, "3": True}
        link.vote(vote)

        result = Result.objects.get(survey=survey)
        self.assertIsInstance(result.vote, dict)
        self.assertEqual(result.vote["0"], 3)
        self.assertEqual(result.vote["3"], True)

        # ActiveLink must be destroyed after voting
        self.assertRaises(ObjectDoesNotExist, ActiveLink.objects.get, survey=survey)

    def test_l2e_survey_creation(self):
        """L2E survey stores questions list and correct kind."""
        questions = ["Statement A", "Statement B"]
        survey = Survey.objects.create(
            question="L2E parent",
            questions=questions,
            active=True,
            kind="L2E",
            expiry=datetime.datetime(2099, 1, 1, 0, 0),
            participants=1,
            voted=0,
        )
        saved = Survey.objects.get(pk=survey.pk)
        self.assertEqual(saved.kind, "L2E")
        self.assertEqual(saved.questions, questions)

    def test_l2e_vote_stored_as_dict(self):
        """Voting on an L2E survey stores a dict in Result.vote."""
        survey = Survey.objects.create(
            question="L2E vote test",
            questions=["S1", "S2"],
            active=True,
            kind="L2E",
            expiry=datetime.datetime(2099, 1, 1, 0, 0),
            participants=1,
            voted=0,
        )
        link = ActiveLink.objects.create(participant=self.participant, survey=survey)
        vote = {"0": 2, "1": 5, "2": False}
        link.vote(vote)

        result = Result.objects.get(survey=survey)
        self.assertIsInstance(result.vote, dict)
        self.assertEqual(result.vote["0"], 2)
        self.assertEqual(result.vote["2"], False)

        # ActiveLink must be destroyed after voting
        self.assertRaises(ObjectDoesNotExist, ActiveLink.objects.get, survey=survey)

    def test_survey_institution_link_counts(self):
        """
        /survey/{id}/institutions/ returns correct link_count and total_count per
        institution with no double-counting when both ActiveLinks and Results exist.

        Setup: 2 institutions, 2 participants each → 4 active links.
        Cast 1 vote per institution (destroys 2 links).
        Expected: each institution has link_count=1, total_count=2.
        """
        inst_a = Institution.objects.create(name="Inst Alpha", country="GB")
        inst_b = Institution.objects.create(name="Inst Beta", country="US")
        disc = Discipline.objects.create(name="TestDisc")

        p_a1 = Participant.objects.create(
            email="a1@test.invalid", name="A1", institution=inst_a, discipline=disc
        )
        p_a2 = Participant.objects.create(
            email="a2@test.invalid", name="A2", institution=inst_a, discipline=disc
        )
        p_b1 = Participant.objects.create(
            email="b1@test.invalid", name="B1", institution=inst_b, discipline=disc
        )
        p_b2 = Participant.objects.create(
            email="b2@test.invalid", name="B2", institution=inst_b, discipline=disc
        )

        survey = Survey.objects.create(
            question="Institution count test survey",
            active=True,
            kind="LI",
            expiry=datetime.datetime(2099, 1, 1, 0, 0),
            participants=4,
            voted=0,
        )

        link_a1 = ActiveLink.objects.create(participant=p_a1, survey=survey)
        ActiveLink.objects.create(participant=p_a2, survey=survey)
        link_b1 = ActiveLink.objects.create(participant=p_b1, survey=survey)
        ActiveLink.objects.create(participant=p_b2, survey=survey)

        # Cast one vote per institution — destroys that ActiveLink, creates a Result
        link_a1.vote(3)
        link_b1.vote(4)

        user = User.objects.create(username="counttest")
        user.set_password("password")
        user.save()
        self.client.login(username="counttest", password="password")

        resp = self.client.get(f"/api/survey/{survey.id}/institutions/")
        self.assertEqual(resp.status_code, 200)

        import json as json_mod

        data = json_mod.loads(resp.content)
        results = {r["name"]: r for r in data["results"]}

        self.assertIn("Inst Alpha", results)
        self.assertIn("Inst Beta", results)

        for name in ("Inst Alpha", "Inst Beta"):
            self.assertEqual(results[name]["link_count"], 1, f"{name} link_count")
            self.assertEqual(results[name]["total_count"], 2, f"{name} total_count")

    def test_invalid_kind_rejected(self):
        """Survey.full_clean() raises ValidationError for an unrecognised kind."""
        from django.core.exceptions import ValidationError as DjangoValidationError

        survey = Survey(
            question="Bad kind test",
            active=True,
            kind="INVALID",
            expiry=datetime.datetime(2099, 1, 1, 0, 0),
        )
        with self.assertRaises(DjangoValidationError):
            survey.full_clean()

    def test_token_validation_valid(self):
        """GET /api/vote/{token}/ returns 200 for an existing (unused) token."""
        link = ActiveLink.objects.get()
        resp = self.client.get(f"/api/vote/{link.unique_link}/")
        self.assertEqual(resp.status_code, 200)
        import json as json_mod

        data = json_mod.loads(resp.content)
        self.assertEqual(data["status"], "valid")

    def test_token_validation_used_or_missing(self):
        """GET /api/vote/{token}/ returns 404 for a missing or already-used token."""
        resp = self.client.get("/api/vote/nonexistent-token-xyz/")
        self.assertEqual(resp.status_code, 404)

        # Use the link and check again
        link = ActiveLink.objects.get()
        uid = link.unique_link
        link.vote(3)
        resp = self.client.get(f"/api/vote/{uid}/")
        self.assertEqual(resp.status_code, 404)

    def _l3c_link(self):
        """Create a fresh L3C survey and return an ActiveLink for it."""
        survey = Survey.objects.create(
            question="L3C validation test",
            questions=["S1", "S2", "S3", "Expertise"],
            active=True,
            kind="L3C",
            expiry=datetime.datetime(2099, 1, 1, 0, 0),
            participants=1,
            voted=0,
        )
        return ActiveLink.objects.create(participant=self.participant, survey=survey)

    def test_vote_validation_rejects_wrong_keys(self):
        """POST /api/vote/ with unexpected dict keys returns 400 and preserves the token."""
        link = self._l3c_link()
        # "expertise" key is no longer valid for L3C; the checkbox key is now "3"
        bad_vote = {"0": 3, "1": 4, "2": 2, "expertise": True}
        response = self.client.post(
            "/api/vote/",
            {"unique_id": str(link.unique_link), "vote": json.dumps(bad_vote)},
        )
        self.assertEqual(response.status_code, 400)
        # Token must survive a rejected vote
        self.assertTrue(
            ActiveLink.objects.filter(unique_link=link.unique_link).exists()
        )

    def test_vote_validation_rejects_out_of_range_likert_in_dict(self):
        """POST /api/vote/ with an out-of-range Likert value inside a dict vote returns 400."""
        link = self._l3c_link()
        bad_vote = {"0": 99, "1": 4, "2": 2, "3": True}  # 99 is outside 1–5
        response = self.client.post(
            "/api/vote/",
            {"unique_id": str(link.unique_link), "vote": json.dumps(bad_vote)},
        )
        self.assertEqual(response.status_code, 400)
        self.assertTrue(
            ActiveLink.objects.filter(unique_link=link.unique_link).exists()
        )

    def test_vote_validation_rejects_wrong_checkbox_type(self):
        """POST /api/vote/ with an integer where a boolean is expected returns 400."""
        link = self._l3c_link()
        bad_vote = {"0": 3, "1": 4, "2": 2, "3": 1}  # 1 is int, not bool
        response = self.client.post(
            "/api/vote/",
            {"unique_id": str(link.unique_link), "vote": json.dumps(bad_vote)},
        )
        self.assertEqual(response.status_code, 400)
        self.assertTrue(
            ActiveLink.objects.filter(unique_link=link.unique_link).exists()
        )

    def test_vote_validation_accepts_valid_dict_vote(self):
        """POST /api/vote/ with a correctly-structured dict vote returns 200 and stores the result."""
        link = self._l3c_link()
        valid_vote = {"0": 3, "1": 4, "2": 2, "3": True}
        response = self.client.post(
            "/api/vote/",
            {"unique_id": str(link.unique_link), "vote": json.dumps(valid_vote)},
        )
        self.assertEqual(response.status_code, 200)
        result = Result.objects.get(survey=link.survey)
        self.assertEqual(result.vote["0"], 3)
        self.assertEqual(result.vote["3"], True)
        self.assertIsInstance(result.vote["3"], bool)
        self.assertFalse(
            ActiveLink.objects.filter(unique_link=link.unique_link).exists()
        )

    def test_vote_validation_rejects_out_of_range_likert(self):
        """POST /api/vote/ with an out-of-range Likert value returns 400 and preserves the token."""
        link = ActiveLink.objects.get(survey=self.survey)
        response = self.client.post(
            "/api/vote/",
            {"unique_id": str(link.unique_link), "vote": "99"},
        )
        self.assertEqual(response.status_code, 400)
        self.assertTrue(
            ActiveLink.objects.filter(unique_link=link.unique_link).exists()
        )


class SurveyTemplateTestCase(TestCase):
    """Tests for SurveyTemplate model and /api/survey/templates/ endpoints."""

    @staticmethod
    def _make_template(label, slug, slots, is_builtin=False):
        """Helper: create a SurveyTemplate with its SurveyTemplateSlot rows."""
        from iasc.models import SurveyTemplate, SurveyTemplateSlot

        tmpl = SurveyTemplate.objects.create(
            label=label, slug=slug, is_builtin=is_builtin
        )
        for order, slot in enumerate(slots):
            SurveyTemplateSlot.objects.create(
                template=tmpl,
                order=order,
                slot_id=slot["id"],
                type=slot["type"],
                placeholder=slot.get("placeholder", ""),
            )
        return tmpl

    def setUp(self):
        super().setUp()
        from iasc.models import SurveyTemplate

        self.SurveyTemplate = SurveyTemplate

        user = User.objects.create(username="tmpl_user")
        user.set_password("password")
        user.save()
        self.client = __import__("django.test", fromlist=["Client"]).Client()
        self.client.login(username="tmpl_user", password="password")

        self.json_mt = "application/json"

        # Create a non-builtin template for mutation tests
        self.custom = self._make_template(
            "Custom Template",
            "CUSTOM",
            [{"id": "q0", "type": "likert", "placeholder": "Rate this"}],
        )

    def _get(self, url, expected_status=200):
        resp = self.client.get(url)
        self.assertEqual(resp.status_code, expected_status)
        return json.loads(resp.content)

    def _post(self, url, data, expected_status=201):
        resp = self.client.post(url, json.dumps(data), content_type="application/json")
        self.assertEqual(resp.status_code, expected_status)
        return json.loads(resp.content)

    def _patch(self, url, data, expected_status=200):
        resp = self.client.patch(url, json.dumps(data), content_type="application/json")
        self.assertEqual(resp.status_code, expected_status)
        return json.loads(resp.content)

    def _delete(self, url, expected_status=204):
        resp = self.client.delete(url)
        self.assertEqual(resp.status_code, expected_status)
        return resp

    # --- list / retrieve ---

    def test_list_returns_seeded_builtins(self):
        """The 4 builtin templates are returned by the list endpoint."""
        data = self._get("/api/survey/templates/")
        slugs = {t["slug"] for t in data}
        self.assertIn("LI", slugs)
        self.assertIn("L2E", slugs)
        self.assertIn("L3C", slugs)
        self.assertIn("LI3", slugs)

    def test_retrieve_by_slug(self):
        data = self._get("/api/survey/templates/LI/")
        self.assertEqual(data["slug"], "LI")
        self.assertTrue(data["is_builtin"])
        self.assertIsInstance(data["slots"], list)
        self.assertGreater(len(data["slots"]), 0)

    def test_survey_serializer_includes_template_slots(self):
        """GET /api/survey/<id>/ includes template_slots from the linked template."""
        survey = Survey.objects.create(
            question="Template slots test",
            active=True,
            kind="LI",
            expiry=datetime.datetime(2099, 1, 1),
        )
        data = self._get(f"/api/survey/{survey.id}/")
        self.assertIn("template_slots", data)
        self.assertIsNotNone(data["template_slots"])
        self.assertEqual(data["template_slots"][0]["type"], "likert")

    # --- create ---

    def test_create_valid_template(self):
        data = self._post(
            "/api/survey/templates/",
            {
                "label": "My New Template",
                "slug": "NEW1",
                "slots": [
                    {"id": "q0", "type": "likert", "placeholder": "Statement 1"},
                    {"id": "q1", "type": "checkbox", "placeholder": "Expertise"},
                ],
            },
        )
        self.assertEqual(data["slug"], "NEW1")
        self.assertEqual(len(data["slots"]), 2)
        self.assertFalse(data["is_builtin"])

    def test_create_duplicate_slug_rejected(self):
        self._post(
            "/api/survey/templates/",
            {
                "label": "Dup",
                "slug": "CUSTOM",
                "slots": [{"id": "q0", "type": "likert"}],
            },
            expected_status=400,
        )

    def test_create_empty_slots_rejected(self):
        self._post(
            "/api/survey/templates/",
            {"label": "Bad", "slug": "BAD1", "slots": []},
            expected_status=400,
        )

    def test_create_invalid_slot_type_rejected(self):
        self._post(
            "/api/survey/templates/",
            {
                "label": "BadType",
                "slug": "BAD2",
                "slots": [{"id": "q0", "type": "radio"}],
            },
            expected_status=400,
        )

    def test_create_duplicate_slot_ids_rejected(self):
        self._post(
            "/api/survey/templates/",
            {
                "label": "DupIds",
                "slug": "BAD3",
                "slots": [
                    {"id": "q0", "type": "likert"},
                    {"id": "q0", "type": "checkbox"},
                ],
            },
            expected_status=400,
        )

    # --- update ---

    def test_patch_custom_template_label(self):
        data = self._patch(
            f"/api/survey/templates/{self.custom.slug}/",
            {"label": "Renamed"},
        )
        self.assertEqual(data["label"], "Renamed")

    def test_patch_slug_is_ignored(self):
        """slug is read-only on update; sending a new value has no effect."""
        data = self._patch(
            f"/api/survey/templates/{self.custom.slug}/",
            {"slug": "NEWSLUG"},
        )
        self.assertEqual(data["slug"], self.custom.slug)
        self.assertFalse(self.SurveyTemplate.objects.filter(slug="NEWSLUG").exists())

    def test_patch_builtin_template_rejected(self):
        self._patch(
            "/api/survey/templates/LI/",
            {"label": "Hacked"},
            expected_status=400,
        )

    def test_patch_slots_blocked_when_surveys_exist(self):
        """Changing slots is blocked if surveys already use this template."""
        Survey.objects.create(
            question="Blocking survey",
            active=True,
            kind=self.custom.slug,
            expiry=datetime.datetime(2099, 1, 1),
        )
        self._patch(
            f"/api/survey/templates/{self.custom.slug}/",
            {"slots": [{"id": "q0", "type": "checkbox", "placeholder": "Changed"}]},
            expected_status=400,
        )

    def test_patch_slots_allowed_when_no_surveys_exist(self):
        """Changing slot structure is allowed when no surveys use this template."""
        data = self._patch(
            f"/api/survey/templates/{self.custom.slug}/",
            {
                "slots": [
                    {"id": "q0", "type": "likert", "placeholder": "New statement"},
                    {"id": "q1", "type": "checkbox", "placeholder": "New checkbox"},
                ]
            },
        )
        self.assertEqual(len(data["slots"]), 2)
        self.assertEqual(data["slots"][1]["type"], "checkbox")

    def test_survey_results_serializer_includes_template_slots(self):
        """GET /api/survey/results/ includes template_slots for each result survey."""
        from iasc.models import Result

        survey = Survey.objects.create(
            question="Result slots test",
            active=True,
            kind="LI",
            expiry=datetime.datetime(2099, 1, 1),
        )
        institution = Institution.objects.create(name="ResultSlotsUni", country="GB")
        discipline = Discipline.objects.create(name="ResultSlotsDiscipline")
        Result.objects.create(
            survey=survey,
            vote=3,
            institution=institution,
            discipline=discipline,
        )
        data = self._get("/api/survey/results/")
        results = data.get("results", [])
        matching = [r for r in results if r["id"] == survey.id]
        self.assertEqual(len(matching), 1)
        self.assertIn("template_slots", matching[0])
        self.assertIsNotNone(matching[0]["template_slots"])
        self.assertEqual(matching[0]["template_slots"][0]["type"], "likert")

    # --- delete ---

    def test_delete_custom_template(self):
        unused = self._make_template(
            "Unused", "UNUSED1", [{"id": "q0", "type": "likert", "placeholder": "x"}]
        )
        self._delete(f"/api/survey/templates/{unused.slug}/")
        self.assertFalse(self.SurveyTemplate.objects.filter(slug="UNUSED1").exists())

    def test_delete_builtin_rejected(self):
        resp = self.client.delete("/api/survey/templates/LI/")
        self.assertEqual(resp.status_code, 400)
        self.assertTrue(self.SurveyTemplate.objects.filter(slug="LI").exists())

    def test_delete_template_with_surveys_rejected(self):
        Survey.objects.create(
            question="Prevent delete",
            active=True,
            kind=self.custom.slug,
            expiry=datetime.datetime(2099, 1, 1),
        )
        resp = self.client.delete(f"/api/survey/templates/{self.custom.slug}/")
        self.assertEqual(resp.status_code, 400)

    # --- survey_count ---

    def test_survey_count_in_template_response(self):
        Survey.objects.create(
            question="Count test",
            active=True,
            kind=self.custom.slug,
            expiry=datetime.datetime(2099, 1, 1),
        )
        data = self._get(f"/api/survey/templates/{self.custom.slug}/")
        self.assertEqual(data["survey_count"], 1)

    # --- validate_survey_kind uses DB ---

    def test_survey_kind_validation_uses_db(self):
        """Survey.full_clean() accepts a slug that exists in SurveyTemplate."""
        survey = Survey(
            question="DB-validated kind",
            active=True,
            kind=self.custom.slug,
            expiry=datetime.datetime(2099, 1, 1),
        )
        survey.full_clean()  # should not raise

    def test_survey_kind_invalid_still_rejected(self):
        from django.core.exceptions import ValidationError as DjangoValidationError

        survey = Survey(
            question="Bad kind",
            active=True,
            kind="NONEXISTENT",
            expiry=datetime.datetime(2099, 1, 1),
        )
        with self.assertRaises(DjangoValidationError):
            survey.full_clean()
